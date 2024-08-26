import openpyxl

# ---------- Step One ----------

# wb = openpyxl.load_workbook("octobre.xlsx")

# print(wb.sheetnames)

# # sheet = wb['Feuil1'] # sheet = wb[wb.sheetnames[0]]
# sheet = wb.active
#
# # cell = sheet['A1']
#
# # row = 2
# # for row in range(2, 7):
# #     cell = sheet.cell(row, column=2)
# #     print(cell.value)
#
# print(sheet.max_row)
# print(sheet.max_column)

# ---------- Step Two ----------
wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)

# {"Pommes": (760, 660, 900) }

def add_data_from_wb(workbook, dictionnary):
    sheet = workbook.active
    for row in range(2,sheet.max_row):
        article_name = sheet.cell(row,1).value
        if not article_name:
            break
        total_sales = sheet.cell(row,4).value
        if dictionnary.get(article_name):
            dictionnary[article_name].append(total_sales)
        else:
            dictionnary[article_name] = [total_sales]

data = {}
add_data_from_wb(wb1, data)
add_data_from_wb(wb2, data)
add_data_from_wb(wb3, data)

print(data)

# ---------- Step Three ----------

output_wb = openpyxl.Workbook()
sheet = output_wb.active
# sheet.title = "Sheet1"

sheet["A1"] = "Article"
sheet["B1"] = "October"
sheet["C1"] = "November"
sheet["D1"] = "December"



row = 2
for i in data.items():
    # print(i)
    article_name = i[0]
    sales = i[1]
    sheet.cell(row, 1).value = article_name
    for j in range(0, len(sales)):
        sheet.cell(row, 2+j).value = sales[j]
    row += 1


output_wb.save("total_quarterly_sales.xlsx")
